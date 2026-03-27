"""
Importa dados de abrangência/pactuação dos Excel para o banco SQLite.

Arquivos fonte (pasta abrangencia/):
  - ITENS x PROCEDIMENTOS.xlsx  -> tabela item_procedimento
  - ABRANGENCIA.xlsx            -> tabela abrangencia
  - REFERENCIA.xlsx             -> tabela referencia

Uso:
  python import_abrangencia.py
  python import_abrangencia.py --reset   # limpa tabelas antes de importar
"""

import os
import sys
import sqlite3
import openpyxl

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "saude_real.db")
ABRANGENCIA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "abrangencia")


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def parse_cod_nome(text):
    """Extrai código e nome de strings como '0201010011 - AMNIOCENTESE'."""
    text = str(text).strip()
    if " - " in text:
        parts = text.split(" - ", 1)
        return parts[0].strip(), parts[1].strip()
    return text, text


def import_itens_procedimentos(conn):
    """Importa ITENS x PROCEDIMENTOS.xlsx -> item_procedimento."""
    filepath = os.path.join(ABRANGENCIA_DIR, "ITENS x PROCEDIMENTOS.xlsx")
    if not os.path.exists(filepath):
        print(f"SKIP: {filepath} não encontrado")
        return 0

    wb = openpyxl.load_workbook(filepath, read_only=True)
    cursor = conn.cursor()
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tipo = sheet_name.strip().upper()
        last_item_cod = ""
        last_item_nome = ""

        for row in ws.iter_rows(min_row=10, values_only=True):
            item_raw = str(row[0]).strip() if row[0] else ""
            proc_raw = str(row[1]).strip() if row[1] else ""

            if item_raw and item_raw != "None":
                last_item_cod, last_item_nome = parse_cod_nome(item_raw)

            if proc_raw and proc_raw != "None" and last_item_cod:
                proc_cod, proc_nome = parse_cod_nome(proc_raw)
                cursor.execute(
                    "INSERT OR IGNORE INTO item_procedimento (tipo, item_cod, item_nome, proc_cod, proc_nome) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (tipo, last_item_cod, last_item_nome, proc_cod, proc_nome),
                )
                total += 1

    wb.close()
    conn.commit()
    print(f"item_procedimento: {total} registros importados")
    return total


def import_abrangencia(conn):
    """Importa ABRANGENCIA.xlsx -> abrangencia."""
    filepath = os.path.join(ABRANGENCIA_DIR, "ABRANGENCIA.xlsx")
    if not os.path.exists(filepath):
        print(f"SKIP: {filepath} não encontrado")
        return 0

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    cursor = conn.cursor()
    total = 0

    for row in ws.iter_rows(min_row=10, values_only=True):
        tipo = str(row[1]).strip() if row[1] else ""
        financiamento = str(row[2]).strip() if row[2] else ""
        item_raw = str(row[3]).strip() if row[3] else ""
        executor = str(row[4]).strip() if row[4] else ""

        if not item_raw or item_raw == "None":
            continue

        item_cod, item_nome = parse_cod_nome(item_raw)

        qtd = row[5] if row[5] else 0
        vl_unit = row[6] if row[6] else 0.0
        vl_total = row[7] if row[7] else 0.0

        try:
            qtd = int(qtd)
        except (ValueError, TypeError):
            qtd = 0
        try:
            vl_unit = float(vl_unit)
        except (ValueError, TypeError):
            vl_unit = 0.0
        try:
            vl_total = float(vl_total)
        except (ValueError, TypeError):
            vl_total = 0.0

        cursor.execute(
            "INSERT INTO abrangencia (tipo, financiamento, item_cod, item_nome, "
            "municipio_executor, quantidade, valor_unitario, valor_total) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (tipo, financiamento, item_cod, item_nome, executor, qtd, vl_unit, vl_total),
        )
        total += 1

    wb.close()
    conn.commit()
    print(f"abrangencia: {total} registros importados")
    return total


def import_referencia(conn):
    """Importa REFERENCIA.xlsx -> referencia."""
    filepath = os.path.join(ABRANGENCIA_DIR, "REFERENCIA.xlsx")
    if not os.path.exists(filepath):
        print(f"SKIP: {filepath} não encontrado")
        return 0

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    cursor = conn.cursor()
    total = 0
    batch = []

    for row in ws.iter_rows(min_row=10, values_only=True):
        tipo = str(row[1]).strip() if row[1] else ""
        financiamento = str(row[2]).strip() if row[2] else ""
        encaminhador = str(row[3]).strip() if row[3] else ""
        item_raw = str(row[4]).strip() if row[4] else ""
        executor = str(row[5]).strip() if row[5] else ""

        if not item_raw or item_raw == "None":
            continue

        item_cod, item_nome = parse_cod_nome(item_raw)

        qtd = row[6] if row[6] else 0
        vl_unit = row[7] if row[7] else 0.0
        vl_total = row[8] if row[8] else 0.0

        try:
            qtd = int(qtd)
        except (ValueError, TypeError):
            qtd = 0
        try:
            vl_unit = float(vl_unit)
        except (ValueError, TypeError):
            vl_unit = 0.0
        try:
            vl_total = float(vl_total)
        except (ValueError, TypeError):
            vl_total = 0.0

        batch.append((tipo, financiamento, encaminhador, item_cod, item_nome,
                       executor, qtd, vl_unit, vl_total))

        if len(batch) >= 5000:
            cursor.executemany(
                "INSERT INTO referencia (tipo, financiamento, municipio_encaminhador, "
                "item_cod, item_nome, municipio_executor, quantidade, valor_unitario, valor_total) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                batch,
            )
            total += len(batch)
            batch = []

    if batch:
        cursor.executemany(
            "INSERT INTO referencia (tipo, financiamento, municipio_encaminhador, "
            "item_cod, item_nome, municipio_executor, quantidade, valor_unitario, valor_total) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            batch,
        )
        total += len(batch)

    wb.close()
    conn.commit()
    print(f"referencia: {total} registros importados")
    return total


def main():
    reset = "--reset" in sys.argv

    from db_manager import create_tables
    create_tables()

    conn = get_connection()

    if reset:
        print("Limpando tabelas de abrangência...")
        conn.execute("DELETE FROM item_procedimento")
        conn.execute("DELETE FROM abrangencia")
        conn.execute("DELETE FROM referencia")
        conn.commit()

    print("Importando dados de abrangência...\n")
    import_itens_procedimentos(conn)
    import_abrangencia(conn)
    import_referencia(conn)

    print("\n=== Resumo ===")
    for table in ["item_procedimento", "abrangencia", "referencia"]:
        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {table}: {count} registros")

    conn.close()
    print("\nImportação concluída.")


if __name__ == "__main__":
    main()
